"""
MSHSAA Fall Softbal - Class & District Assignment Scraper
============================================================
Scrapes all class/district assignments from 2012-2013 through 2025-2026.
 
Usage:
    python mshsaa_fall_softball_scraper.py
 
Output:
    fall_softball_districts.xlsx  (sheet "ALL TEAMS": Team Name, Class, District, Season)
 
Name formatting logic:
    Each MSHSAA page contains a flat school roster where every entry looks like:
        "246 Bucklin Logo Bucklin District 12 ..."
        "392 Northeast (Cairo) Logo Northeast (Cairo) District 10 ..."
 
    The number is the school's unique MSHSAA ID. The name between the ID and
    the word "Logo" is always the clean canonical school name with no sub-notes.
 
    District section anchors link to Schedule.aspx?s=<ID>, so we can look up
    the canonical name by ID and compare it to the district display text:
        - display == canonical              -> keep as-is
        - display == canonical + "(note)"  -> "canonical with note"
 
    Examples:
        ID 246  canonical "Bucklin"                    display "Bucklin (Macon County)"
                -> "Bucklin with Macon County"
        ID 392  canonical "Northeast (Cairo)"          display "Northeast (Cairo)"
                -> "Northeast (Cairo)"   (exact match, kept as-is)
        ID 454  canonical "Southwest (Livingston County)"
                                                       display "Southwest (Livingston County) (Breckenridge)"
                -> "Southwest (Livingston County) with Breckenridge"
 
Requirements:
    pip install requests beautifulsoup4 lxml openpyxl
"""
 
import re
import time
import logging
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
 
logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)
 
# ── constants ────────────────────────────────────────────────────────────────
 
BASE_URL    = "https://www.mshsaa.org/Activities/ClassAndDistrictAssignments.aspx"
ACTIVITY    = 38
NUM_CLASSES = 5
 
SEASONS = [
    ("2025-2026", None),
    ("2024-2025", 2024),
    ("2023-2024", 2023),
    ("2022-2023", 2022),
    ("2021-2022", 2021),
    ("2020-2021", 2020),
    ("2019-2020", 2019),
    ("2018-2019", 2018),
    ("2017-2018", 2017),
    ("2016-2017", 2016),
    ("2015-2016", 2015),
    ("2014-2015", 2014),
    ("2013-2014", 2013),
    ("2012-2013", 2012),
]
 
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Referer": "https://www.mshsaa.org/",
}
 
REQUEST_DELAY = 1.5
MAX_RETRIES   = 3
 
# ── fetch ────────────────────────────────────────────────────────────────────
 
def fetch(url, params):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.get(url, params=params, headers=HEADERS, timeout=20)
            r.raise_for_status()
            return BeautifulSoup(r.text, "lxml")
        except Exception as exc:
            log.warning("  Attempt %d failed: %s", attempt, exc)
            if attempt < MAX_RETRIES:
                time.sleep(3)
    raise RuntimeError(f"Failed after {MAX_RETRIES} attempts: {url} {params}")
 
# ── canonical name map (ID -> name) ──────────────────────────────────────────
 
def build_id_to_canonical(soup):
    """
    Parse the flat school roster to build a dict of {school_id: canonical_name}.
 
    Each roster entry in the page's plain text looks like:
        "246 Bucklin Logo Bucklin District 12 12 39.795 -92.879 fas fa-location-pin"
        "392 Northeast (Cairo) Logo Northeast (Cairo) District 10 10 ..."
 
    We extract the integer ID and the name that appears BEFORE the word "Logo".
    That pre-Logo name is always the clean canonical school name — MSHSAA only
    appends location/co-op notes in the district section display text, not here.
    """
    raw = soup.get_text(" ", strip=True)
 
    # Anchor the search to the roster block
    start = raw.find("School ID School District")
    if start == -1:
        start = 0
 
    # Match: <digits> <name> Logo
    # The name can contain letters, digits, spaces, parens, periods, hyphens, apostrophes
    pattern = re.compile(r'\b(\d+)\s+([\w][\w\s().,-]+?)\s+Logo\b')
 
    id_map = {}
    for m in pattern.finditer(raw[start:]):
        school_id = int(m.group(1))
        name = re.sub(r'\s+', ' ', m.group(2).strip())
        if name and len(name) >= 2:
            id_map[school_id] = name
 
    return id_map
 
# ── name formatting ───────────────────────────────────────────────────────────
 
def strip_outer_parens(s):
    """Remove outermost parens: '(Macon County)' -> 'Macon County'.
    Handles nested: '(Jefferson (Conception))' -> 'Jefferson (Conception)'."""
    s = s.strip()
    if not (s.startswith("(") and s.endswith(")")):
        return s
    depth = 0
    for i, ch in enumerate(s):
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        if depth == 0 and i < len(s) - 1:
            # Outer paren closed before end — parens don't wrap everything
            return s
    return s[1:-1]
 
 
def format_name(school_id, display_text, id_map):
    """
    Return the correctly formatted team name using the ID-based canonical lookup.
 
      - If display matches canonical exactly -> return as-is
      - If display == canonical + " (note)"  -> return "canonical with note"
      - Otherwise                            -> return display unchanged
    """
    display = display_text.strip()
    canonical = id_map.get(school_id)
 
    if canonical is None:
        log.debug("  No canonical found for ID %d (%r)", school_id, display)
        return display
 
    if display == canonical:
        return display
 
    if display.startswith(canonical):
        remainder = display[len(canonical):].strip()
        if remainder.startswith("("):
            inner = strip_outer_parens(remainder)
            return f"{canonical} with {inner}"
 
    # Display doesn't start with canonical — return as-is and log for review
    log.debug("  Unexpected format ID=%d canonical=%r display=%r", school_id, canonical, display)
    return display
 
# ── scrape one class page ─────────────────────────────────────────────────────
 
def scrape_class(class_num, year_param, season_label):
    params = {"alg": ACTIVITY, "class": class_num}
    if year_param is not None:
        params["year"] = year_param
 
    log.info("  Fetching Class %d  season %s", class_num, season_label)
    soup = fetch(BASE_URL, params)
 
    # Build ID -> canonical name from the flat roster
    id_map = build_id_to_canonical(soup)
    log.debug("    ID map size: %d", len(id_map))
 
    records = []
 
    # Each district is headed by an <h4> tag like "District 1"
    for h4 in soup.find_all("h4"):
        m = re.match(r"District\s+(\d+)", h4.get_text(strip=True))
        if not m:
            continue
        district_num = int(m.group(1))
 
        # Find the <ul> of schools that follows this <h4>
        school_ul = None
        sibling = h4.find_next_sibling()
        while sibling:
            if sibling.name == "ul":
                school_ul = sibling
                break
            if sibling.name == "h4":
                break
            sibling = sibling.find_next_sibling()
 
        if not school_ul:
            log.warning("    No <ul> after District %d", district_num)
            continue
 
        for li in school_ul.find_all("li", recursive=False):
            a = li.find("a", href=re.compile(r"Schedule\.aspx"))
            if not a:
                continue
 
            # Extract school ID from href: "Schedule.aspx?s=246&alg=5"
            href = a.get("href", "")
            id_match = re.search(r"[?&]s=(\d+)", href)
            if not id_match:
                continue
            school_id = int(id_match.group(1))
 
            raw_name = a.get_text(" ", strip=True).replace(" Host", "").strip()
            formatted = format_name(school_id, raw_name, id_map)
 
            records.append({
                "Team Name": formatted,
                "Class":     class_num,
                "District":  district_num,
                "Season":    season_label,
            })
 
    log.info("    -> %d teams", len(records))
    return records
 
# ── scrape all seasons ────────────────────────────────────────────────────────
 
def scrape_all():
    all_records = []
    for season_label, year_param in SEASONS:
        log.info("Season: %s", season_label)
        for class_num in range(1, NUM_CLASSES + 1):
            try:
                records = scrape_class(class_num, year_param, season_label)
                all_records.extend(records)
            except Exception as exc:
                log.error("  ERROR class %d season %s: %s", class_num, season_label, exc)
            time.sleep(REQUEST_DELAY)
    return all_records
 
# ── Excel output ──────────────────────────────────────────────────────────────
 
def write_excel(records, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "ALL TEAMS"
 
    header_font  = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", start_color="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center")
 
    col_headers = ["Team Name", "Class", "District", "Season"]
    col_widths  = [45, 8, 10, 12]
 
    for col, (h, w) in enumerate(zip(col_headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w
 
    ws.row_dimensions[1].height = 18
    fill_even = PatternFill("solid", start_color="DCE6F1")
    data_font = Font(name="Arial", size=10)
 
    for row_num, rec in enumerate(records, start=2):
        ws.cell(row=row_num, column=1, value=rec["Team Name"]).font = data_font
        ws.cell(row=row_num, column=2, value=rec["Class"]).font     = data_font
        ws.cell(row=row_num, column=3, value=rec["District"]).font  = data_font
        ws.cell(row=row_num, column=4, value=rec["Season"]).font    = data_font
 
        if row_num % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row_num, column=col).fill = fill_even
 
        for col in (2, 3):
            ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="center")
 
    ws.freeze_panes = "A2"
    wb.save(output_path)
    log.info("Saved %d records to %s", len(records), output_path)
 
# ── entry point ───────────────────────────────────────────────────────────────
 
if __name__ == "__main__":
    output = "fall_softball_districts.xlsx"
    log.info("Starting MSHSAA Fall Softball scraper")
    log.info("%d seasons x %d classes = %d pages",
             len(SEASONS), NUM_CLASSES, len(SEASONS) * NUM_CLASSES)
    records = scrape_all()
    log.info("Total records: %d", len(records))
    write_excel(records, output)
    log.info("Done.")
