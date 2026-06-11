"""
MSHSAA Boys Basketball - Class & District Assignment Scraper
============================================================
Scrapes all class/district assignments from 2012-2013 through 2025-2026.
 
Usage:
    python mshsaa_bb_scraper.py
 
Output:
    boys_basketball_districts.xlsx  (one sheet "ALL TEAMS", columns: Team Name, Class, District, Season)
 
Name formatting logic:
    - If a school's parenthetical suffix IS part of its official name (e.g. "Northeast (Cairo)"),
      it appears that way in the flat school roster at the top of the page → kept as-is.
    - If a district listing appends extra parenthetical info not in the official name
      (e.g. "Bucklin (Macon County)"), the suffix is a location/co-op note → formatted as
      "Bucklin with Macon County".
 
Requirements:
    pip install requests beautifulsoup4 lxml openpyxl
"""
 
import re
import time
import logging
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
 
logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)
 
# ── constants ────────────────────────────────────────────────────────────────
 
BASE_URL = "https://www.mshsaa.org/Activities/ClassAndDistrictAssignments.aspx"
ACTIVITY  = 5          # Boys Basketball
NUM_CLASSES = 6
 
# Season list: each tuple is (display_label, year_param)
# year_param is the ?year= value; None means use no year param (current season)
SEASONS = [
    ("2025-2026", None),   # current – no year param needed
    ("2024-2025", 2024),
    ("2023-2024", 2023),
    ("2022-2023", 2022),
    ("2021-2022", 2021),
    ("2020-2021", 2020),
    ("2019-2020", 2019),
    ("2018-2019", 2018),
    ("2017-2018", 2017),
    ("2016-2017", 2016),
    ("2015-2016", 2015),
    ("2014-2015", 2014),
    ("2013-2014", 2013),
    ("2012-2013", 2012),
]
 
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Referer": "https://www.mshsaa.org/",
}
 
REQUEST_DELAY = 1.5   # seconds between requests – be polite to MSHSAA servers
MAX_RETRIES   = 3
 
# ── helpers ──────────────────────────────────────────────────────────────────
 
def fetch(url: str, params: dict) -> BeautifulSoup:
    """Fetch a page with retries; return BeautifulSoup."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.get(url, params=params, headers=HEADERS, timeout=20)
            r.raise_for_status()
            return BeautifulSoup(r.text, "lxml")
        except Exception as exc:
            log.warning("  Attempt %d failed: %s", attempt, exc)
            if attempt < MAX_RETRIES:
                time.sleep(3)
    raise RuntimeError(f"Failed to fetch {url} with {params} after {MAX_RETRIES} attempts")
 
 
def build_params(class_num: int, year_param) -> dict:
    p = {"alg": ACTIVITY, "class": class_num}
    if year_param is not None:
        p["year"] = year_param
    return p
 
 
# ── name-formatting logic ────────────────────────────────────────────────────
 
def parse_canonical_names(soup: BeautifulSoup) -> set:
    """
    Parse the flat school roster (the hidden data list at the top of the page)
    to collect all *official* school names, including any parentheticals that
    are genuinely part of the name (e.g. "Northeast (Cairo)").
 
    The roster appears as a <ul> (or inline text block) where each school entry
    looks like:
        <li>27 Atlanta Logo Atlanta District 11 ...</li>
    or as plain text data embedded in the page's school-list section.
 
    The most reliable signal: school names in the "School ID School District …"
    block are separated from metadata by the word "Logo". We extract the token
    immediately after "Logo" (which is the canonical display name).
    """
    canonical = set()
 
    # Strategy 1 – look for the hidden list items that contain "Logo" keyword
    # (MSHSAA renders them as "ID  Name  Logo  Name  District  N  N  lat  lng  icon")
    raw_text = soup.get_text(" ", strip=True)
 
    # Locate the school-id/roster block – it starts after "School ID School District"
    roster_start = raw_text.find("School ID School District")
    if roster_start == -1:
        roster_start = 0
 
    # Find the first district heading to know where the roster ends
    dist_h4 = soup.find("h4")
    roster_end = raw_text.find("District 1", roster_start + 30) if not dist_h4 else len(raw_text)
 
    # Extract all tokens between "Logo" markers
    # Pattern: "<name_before_logo> Logo <canonical_name> District <n>"
    logo_pattern = re.compile(
        r"Logo\s+"                       # the literal word "Logo"
        r"((?:[A-Za-z0-9().,' \-]"       # name: starts with word char or special
        r"(?:[^L]|L(?!ogo))*?)"          # continues until next Logo or District
        r")\s+District\b",
        re.DOTALL
    )
    for m in logo_pattern.finditer(raw_text[roster_start:]):
        name = m.group(1).strip()
        # Clean up stray whitespace
        name = re.sub(r'\s+', ' ', name)
        if name:
            canonical.add(name)
 
    # Strategy 2 – anchor tags linking to school schedule pages contain the
    # display name directly in the <img alt> or in the anchor text
    for a in soup.select("a[href*='Schedule.aspx']"):
        # The anchor text is often "Name Host" or just "Name"
        name = a.get_text(" ", strip=True).replace(" Host", "").strip()
        if name:
            # This gives us the district-section name (may already have extra parens)
            # We store it anyway to cross-reference
            canonical.add(name)
 
    return canonical
 
 
def format_name(district_text: str, canonical_names: set) -> str:
    """
    Given the raw display name from a district listing, return the properly
    formatted team name:
 
      - If the text matches a canonical name exactly → return as-is.
      - If the text is a canonical name with trailing parenthetical info
        that is NOT part of the canonical name → replace trailing "(X)" with
        " with X".
 
    Examples:
        "Northeast (Cairo)"                    → canonical exact match → "Northeast (Cairo)"
        "Bucklin (Macon County)"               → canonical "Bucklin" + note "(Macon County)"
                                                 → "Bucklin with Macon County"
        "Southwest (Livingston County) (Breckenridge)"
                                               → canonical "Southwest (Livingston County)"
                                                 + note "(Breckenridge)"
                                                 → "Southwest (Livingston County) with Breckenridge"
        "Hale (Bosworth)"                      → canonical "Hale" + note "(Bosworth)"
                                                 → "Hale with Bosworth"
        "South Nodaway (Jefferson (Conception))"
                                               → canonical "South Nodaway"
                                                 + note "(Jefferson (Conception))"
                                                 → "South Nodaway with Jefferson (Conception)"
    """
    text = district_text.strip()
 
    # Exact match → done
    if text in canonical_names:
        return text
 
    # Try to find the longest canonical prefix that matches the start of text
    # then confirm the remainder is a parenthetical block
    best_base = None
    best_suffix = None
    for cname in canonical_names:
        if text.startswith(cname):
            remainder = text[len(cname):].strip()
            # remainder must start with '(' and the canonical name must be real
            if remainder.startswith("(") and (best_base is None or len(cname) > len(best_base)):
                best_base = cname
                best_suffix = remainder
 
    if best_base is not None and best_suffix is not None:
        # Strip outer parens from the suffix
        # Handle nested parens like "(Jefferson (Conception))"
        inner = strip_outer_parens(best_suffix)
        return f"{best_base} with {inner}"
 
    # Fallback: no canonical match found – return the text unchanged and log it
    log.debug("  No canonical match for: %r", text)
    return text
 
 
def strip_outer_parens(s: str) -> str:
    """
    Remove the outermost parentheses from a string like "(Macon County)" → "Macon County".
    Handles nested parens: "(Jefferson (Conception))" → "Jefferson (Conception)".
    """
    s = s.strip()
    if s.startswith("(") and s.endswith(")"):
        # Verify the opening paren actually closes at the last char
        depth = 0
        for i, ch in enumerate(s):
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            if depth == 0 and i < len(s) - 1:
                # Closed before end – outer parens don't wrap everything
                return s
        return s[1:-1]
    return s
 
 
# ── main scraping logic ──────────────────────────────────────────────────────
 
def scrape_class(class_num: int, year_param, season_label: str) -> list[dict]:
    """
    Scrape one class page and return list of dicts:
      {"Team Name": ..., "Class": ..., "District": ..., "Season": ...}
    """
    params = build_params(class_num, year_param)
    log.info("  Fetching Class %d  season %s  (params=%s)", class_num, season_label, params)
 
    soup = fetch(BASE_URL, params)
 
    # ── Step 1: build canonical name set from the flat roster ────────────────
    canonical_names = parse_canonical_names(soup)
    log.debug("    Canonical names found: %d", len(canonical_names))
 
    # ── Step 2: parse district sections ──────────────────────────────────────
    records = []
 
    # District sections are <h4> tags like "District 1"
    district_headers = soup.find_all("h4")
 
    for h4 in district_headers:
        header_text = h4.get_text(strip=True)
        m = re.match(r"District\s+(\d+)", header_text)
        if not m:
            continue
        district_num = int(m.group(1))
 
        # The school list follows the h4 – find the next <ul>
        # Walk siblings until we find a <ul> or another <h4>
        sibling = h4.find_next_sibling()
        school_ul = None
        while sibling:
            if sibling.name == "ul":
                school_ul = sibling
                break
            if sibling.name == "h4":
                break
            sibling = sibling.find_next_sibling()
 
        if school_ul is None:
            log.warning("    No <ul> found after %r", header_text)
            continue
 
        for li in school_ul.find_all("li", recursive=False):
            # Find the anchor linking to the school schedule
            a = li.find("a", href=re.compile(r"Schedule\.aspx"))
            if not a:
                continue
            raw_name = a.get_text(" ", strip=True).replace(" Host", "").strip()
            formatted = format_name(raw_name, canonical_names)
            records.append({
                "Team Name": formatted,
                "Class":     class_num,
                "District":  district_num,
                "Season":    season_label,
            })
 
    log.info("    → %d teams", len(records))
    return records
 
 
def scrape_all() -> list[dict]:
    all_records = []
    for season_label, year_param in SEASONS:
        log.info("Season: %s", season_label)
        for class_num in range(1, NUM_CLASSES + 1):
            try:
                records = scrape_class(class_num, year_param, season_label)
                all_records.extend(records)
            except Exception as exc:
                log.error("  ERROR scraping class %d season %s: %s", class_num, season_label, exc)
            time.sleep(REQUEST_DELAY)
    return all_records
 
 
# ── Excel output ─────────────────────────────────────────────────────────────
 
def write_excel(records: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "ALL TEAMS"
 
    # Header row styling (match the sample file style)
    header_font   = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill   = PatternFill("solid", start_color="1F4E79")   # dark blue
    header_align  = Alignment(horizontal="center", vertical="center")
 
    headers = ["Team Name", "Class", "District", "Season"]
    col_widths = [40, 8, 10, 12]
 
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w
 
    ws.row_dimensions[1].height = 18
 
    # Data rows – alternate light shading for readability
    fill_even = PatternFill("solid", start_color="DCE6F1")   # light blue
    data_font = Font(name="Arial", size=10)
    row_num = 2
 
    for rec in records:
        ws.cell(row=row_num, column=1, value=rec["Team Name"]).font = data_font
        ws.cell(row=row_num, column=2, value=rec["Class"]).font     = data_font
        ws.cell(row=row_num, column=3, value=rec["District"]).font  = data_font
        ws.cell(row=row_num, column=4, value=rec["Season"]).font    = data_font
 
        if row_num % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row_num, column=col).fill = fill_even
 
        # Center numeric columns
        for col in (2, 3):
            ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="center")
 
        row_num += 1
 
    # Freeze the header row
    ws.freeze_panes = "A2"
 
    wb.save(output_path)
    log.info("Saved %d records to %s", len(records), output_path)
 
 
# ── entry point ──────────────────────────────────────────────────────────────
 
if __name__ == "__main__":
    output = "boys_basketball_districts.xlsx"
    log.info("Starting MSHSAA Boys Basketball scraper")
    log.info("Seasons: %d  |  Classes: %d  |  Pages: %d",
             len(SEASONS), NUM_CLASSES, len(SEASONS) * NUM_CLASSES)
 
    records = scrape_all()
    log.info("Total records collected: %d", len(records))
    write_excel(records, output)
    log.info("Done.")
